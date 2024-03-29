import customtkinter
from PIL import Image, ImageTk
import openpyxl
from datetime import datetime
import pyautogui
import webbrowser
from urllib.parse import quote
from time import sleep
from cadastro import cadastro

class tela_de_agendamento:
    def __init__(self):
        pass
    
    def main(self, janela_anterior):
        self.click_sair(janela_anterior)
        janela_agendamento = customtkinter.CTk()
        janela_agendamento.geometry('500x300')
        janela_agendamento.title('Agendamento')
        botao_agendamento = customtkinter.CTkButton(janela_agendamento, text = 'Agendar', width=100, height=100, command=self.click_agendar)
        botao_agendamento.place(relx = 0.075, rely = 0.15)
        
        
        botao_avisar = customtkinter.CTkButton(janela_agendamento, text = 'Avisar clientes', width=100, height=100, command=self.tela_aviso)
        botao_avisar.place(relx = 0.390, rely = 0.15)
        
        botao_ver_lista_do_dia = customtkinter.CTkButton(janela_agendamento, text= 'Agendados\ndo dia', width=100, height=100, command=self.click_agendados_do_dia,anchor='center')
        botao_ver_lista_do_dia.place(relx = 0.725, rely = 0.15)
        
        botao_busca_agendamento = customtkinter.CTkButton(janela_agendamento, text ='Buscar\nagendamentos', width=100, height=100, command=self.click_buscar_agendamentos)
        botao_busca_agendamento.place(relx = 0.075, rely = 0.60)
        
        botao_sair = customtkinter.CTkButton(janela_agendamento, text='Sair', width=50, height=10, fg_color='red',command=lambda:self.click_sair(janela_agendamento))
        botao_sair.place(relx = 0.825, rely = 0.03)
        
        botao_excluir = customtkinter.CTkButton(janela_agendamento, text ='Excluir\nagendamento', width=100, height=100, command=self.click_excluir)
        botao_excluir.place(relx = 0.390, rely = 0.60)
        
        #botao_cadastrar_cliente = customtkinter.CTkButton(janela_agendamento, text="Cadastrar",width=200, height=50,command=cadastrar.click_cadastro)
        #botao_cadastrar_cliente.place(relx = 0.725, rely = 0.30)
        
        janela_agendamento.mainloop()
    
    def corrige_telefone(self, telefone):
        codigo_brasil = '55'
        telefone = codigo_brasil + telefone
        return telefone
    def click_agendar(self):
        def agendar():
            nome = str(campo_nome.get())
            data = str(campo_data.get())
            horario = str(campo_horario.get())
            telefone = str(campo_telefone.get())
            print(nome, telefone, horario, data)
            
            try:
                workbook = openpyxl.load_workbook('agendamentos.xlsx')
                planilha = workbook['Sheet1']
                proxima_linha = planilha.max_row + 1
                planilha[f'A{proxima_linha}'] = nome
                planilha[f'B{proxima_linha}'] = data
                planilha[f'C{proxima_linha}'] = horario
                planilha[f'D{proxima_linha}'] = self.corrige_telefone(telefone)
                
            
                workbook.save('agendamentos.xlsx')
                texto = customtkinter.CTkLabel(janela_agendar,text="Agendamento realizado com sucesso",text_color="yellow")
                texto.pack(padx = 10, pady = 10)
            except Exception as e:
                self.tela_erro(e)                            
            
        janela_agendar = customtkinter.CTk()
        janela_agendar.geometry('500x300')
        
        campo_nome = customtkinter.CTkEntry(janela_agendar, placeholder_text= 'Nome do cliente')
        campo_nome.pack(padx = 10, pady = 10)
        
        campo_data = customtkinter.CTkEntry(janela_agendar, placeholder_text= 'Data')
        campo_data.pack(padx = 10, pady = 10)
        
        campo_horario = customtkinter.CTkEntry(janela_agendar, placeholder_text= 'Horário')
        campo_horario.pack(padx = 10, pady = 10)
        
        campo_telefone = customtkinter.CTkEntry(janela_agendar, placeholder_text= 'Telefone')
        campo_telefone.pack(padx = 10, pady = 10)
        
        botao_agendar = customtkinter.CTkButton(janela_agendar, text='agendar', command=agendar)
        botao_agendar.pack(padx = 10, pady = 10)
        
       
        
        janela_agendar.mainloop()
        
    def click_buscar_agendamentos(self):
        def buscar():
            encontrado = False
            nome_digitado_maiusculo = campo_nome.get().upper()

            workbook = openpyxl.load_workbook('agendamentos.xlsx')
            planilha = workbook['Sheet1']
            for linha in planilha.iter_rows(min_row=2, values_only=True):
                nome_planilha_split = linha[0].split(' ')
                nome_planilha = linha[0]
                
                if nome_digitado_maiusculo == nome_planilha_split[0].upper():
                    nome = self.capitalizar_nome(nome_planilha)
                    data = str(linha[1])
                    horario = str(linha[2])
                    telefone = str(linha[3])
                    
                    info_agendamento = nome + '\t' + data + ' '+ horario + '\t' + telefone
                    
                    informacoes = customtkinter.CTkLabel(janela, text=info_agendamento)
                    informacoes.pack(padx = 10, pady = 10)
                    encontrado = True
            if not encontrado:
                texto = customtkinter.CTkLabel(janela,text="Nenhum agendamento encontrado", text_color="yellow")
                texto.pack(padx = 10, pady = 10)

        janela = customtkinter.CTk()
        janela.geometry('500x300')        
        campo_nome = customtkinter.CTkEntry(janela, placeholder_text='Nome')
        campo_nome.pack(padx = 10, pady = 10)
        botao_procurar = customtkinter.CTkButton(janela, text='Buscar',command=buscar)
        botao_procurar.pack(padx = 10, pady = 10)
    
        janela.mainloop()
    
    def click_agendados_do_dia(self):
        encontrado = 0
        janela = customtkinter.CTk()
        janela.geometry("500x300")
        janela.title("Lista")
        
        data_de_hoje = datetime.now()
        data_de_hoje = str(data_de_hoje.strftime('%d/%m/%Y'))
        
        workbook = openpyxl.load_workbook('agendamentos.xlsx')
        planilha = workbook['Sheet1']
        
        cabecalho = customtkinter.CTkLabel(janela, text='Lista de agendados do dia {}:'.format(data_de_hoje))
        cabecalho.pack(padx = 10, pady = 10)
        
        label = customtkinter.CTkScrollableFrame(janela, width=480, height=280)
        label.pack(padx = 10, pady = 10)
        
        cabecalho1 = customtkinter.CTkLabel(label, text='Nome\t\t Data \t\t Horário\t\t Telefone')
        cabecalho1.pack(padx = 10, pady = 10)
        
        for linha in planilha.iter_rows(min_row=2, values_only=True):
            if linha[1] == data_de_hoje:
                encontrado += 1
                dados = self.capitalizar_nome(str(linha[0])) + '\t' + str(linha[1]) + ' ' + str(linha[2]) + '\t' + str(linha[3])
                cliente = customtkinter.CTkLabel(label, text=dados)
                cliente.pack(padx = 10, pady = 10)
        if encontrado == 0:
            texto_sem_agendamentos = customtkinter.CTkLabel(label, text='Sem agendamentos para hoje')
            texto_sem_agendamentos.pack(padx = 10, pady = 10)
        janela.mainloop()
    
    def capitalizar_nome(self, nome):
        nome_separado=nome.split(' ')
        i = 0
        for coluna in nome_separado:
            if i == 0:
                nome_capitalizado = coluna.capitalize()
            else:
                nome_capitalizado = nome_capitalizado + ' ' + coluna.capitalize()
            i += 1
        return str(nome_capitalizado)
    
    def click_avisar(self,janela_anterior):
        self.click_sair(janela_anterior)
        workbook = openpyxl.load_workbook('agendamentos.xlsx')
        planilha = workbook['Sheet1']
     
        data_de_hoje = datetime.now()
        data_de_hoje = str(data_de_hoje.strftime('%d/%m/%Y'))

        
        for linha in planilha.iter_rows(min_row=2, values_only=True):
            nome = self.capitalizar_nome(str(linha[0]))
            primeiro_nome = nome.split(' ')
            primeiro_nome = primeiro_nome[0]
            if linha[1] == data_de_hoje:
                self.bot_wpp(primeiro_nome, str(linha[2]), str(linha[3]))
    
    
    def bot_wpp(self, nome, horario_agendamento, telefone):

        try:
            agora = datetime.now()
            hora = agora.hour
            dia = agora.date()
            dia_formatado = str(dia)
        
            if hora < 12:
                mensagem = 'Bom dia, {}!\nGostaríamos de lembrar de seu horário agendado para hoje às {}'.format(nome, horario_agendamento)
            else:
                mensagem = 'Boa tarde, {}!\nGostaríamos de lembrar de seu horário agendado para hoje às {}'.format(nome, horario_agendamento)
            link = f'https://web.whatsapp.com/send?phone={telefone}&text={quote(mensagem)}'
            webbrowser.open(link)
            sleep(15)
            seta = pyautogui.locateCenterOnScreen('seta.png')
            sleep(2)
            pyautogui.click(seta[0], seta[1])
            sleep(2)
            pyautogui.hotkey('ctrl','w')
            self.operacao_concluida()
        except Exception as e:
            self.tela_erro(e)
        
    def click_excluir(self):
        janela = customtkinter.CTk()
        janela.geometry('500x300')        
        campo_nome = customtkinter.CTkEntry(janela, placeholder_text='Excluir')
        campo_nome.pack(padx = 10, pady = 10)
        botao_procurar = customtkinter.CTkButton(janela, text='Excluir', command=lambda:self.excluir(str(campo_nome.get()),janela))
        botao_procurar.pack(padx = 10, pady = 10)
        janela.mainloop()
    
    def excluir(self, nome_digitado, janela):
        try:
            workbook = openpyxl.load_workbook('agendamentos.xlsx')
            planilha = workbook['Sheet1']
        except Exception as e:
            self.tela_erro(e)
        encontrado = False
        
        row = 2
        for linha in planilha.iter_rows(min_row=2, values_only=True):
            nome_planilha = str(linha[0]).upper()
            nome_digitado = nome_digitado.upper()
            print(nome_planilha)
            print(nome_digitado)
            if nome_planilha == nome_digitado:
                print('Aqui')
                planilha.delete_rows(row)
                encontrado = True
            row += 1
        if not encontrado:
            texto = customtkinter.CTkLabel(janela,text='Não foram encontrados clientes', text_color='yellow')
            texto.pack(padx = 10, pady = 10)
        workbook.save('agendamentos.xlsx')
        
    def tela_aviso(self):
        janela = customtkinter.CTk()
        janela.geometry("600x200")
        janela.title("Aviso")
        
        texto_atencao = customtkinter.CTkLabel(janela, text='ATENÇÃO', bg_color='transparent', text_color="yellow")
        texto_atencao.pack(padx = 10, pady = 10)
        texto = customtkinter.CTkLabel(janela, text='Esta opção controla o mouse e o teclado de seu computador\nPor isso, não use-os enquanto a ação não terminar ou poderá acarretar erros no programa\nCertifique-se ainda de que seu whatsapp web esteja logado neste computador')
        texto.pack(padx = 10, pady = 10)
        
        botao_continuar = customtkinter.CTkButton(janela, text='Continuar', command=lambda:self.click_avisar(janela))
        botao_continuar.pack(padx = 10, pady = 30)
        janela.mainloop()
    
    def click_sair(self, janela):
        janela.destroy()
    
    def limpar_janela(self,janela):
        for widget in janela.winfo_children():
            widget.destroy()

    def operacao_concluida(self):
        janela = customtkinter.CTk()
        janela.geometry('300x100')
        janela.title('Aviso')
        
        texto = customtkinter.CTkLabel(janela, text="Operação concluída com sucesso", text_color="yellow")
        texto.pack(anchor = 'center',fill='both', expand = True)
        
        janela.mainloop()
    
    def tela_erro(self, e):
        janela = customtkinter.CTk()
        janela.geometry('300x100')
        janela.title('Erro')
        
        texto = customtkinter.CTkLabel(janela, text="Operação não concluída, erro inesperado:\n {}".format(e), text_color="yellow")
        texto.pack(anchor = 'center',fill='both', expand = True)
        
        janela.mainloop() 
            
        